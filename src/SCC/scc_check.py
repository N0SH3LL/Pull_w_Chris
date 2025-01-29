"""This is for making sure the SCC itself has required stuff (SCM doc, date, etc.). Checks SCC for required elements. Handles file validation, pattern matching, and date checks.

Functions:
    process_scc_file: Main function to analyze SCC file
    read_excel: Safely load Excel workbook
    find_value_with_regex: Search cells for regex pattern
    check_column_presence: Validate column existence
"""

import openpyxl
import re
import argparse
import os
from datetime import datetime
from typing import Optional, Dict, Any, Union, Pattern

def read_excel(file_path: str) -> Optional[openpyxl.Workbook]:
    """Load an Excel workbook SCC form the specified file path

    Args:
        file_path: Path to Excel file

    Returns:
        Workbook if successful, None if permission denied
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook
    except PermissionError:
        print(f"Permission denied: Unable to access '{file_path}'.") # handles files that are already open somewhere else
        return None

def find_value_with_regex(sheet: openpyxl.worksheet.worksheet.Worksheet,  pattern: Union[str, Pattern], max_diff: int = 5, max_rows: int = 150, max_cols: int = 50) -> Union[str, bool]:
    """Find first cell matching regex pattern in the specified Excel sheet

    Args:
        sheet: Excel worksheet to search
        pattern: Regex pattern to match
        max_diff: Maximum character difference allowed
        max_rows/max_cols: Search limits, to keep it from going for forever and a day

    Returns:
        Matched string or False if not found
    """
    if isinstance(pattern, str):
        compiled_pattern = re.compile(pattern, re.IGNORECASE)
    else:
        compiled_pattern = pattern  

    for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        for cell in row:
            if cell and regex_match_with_diff(compiled_pattern, str(cell), max_diff):
                return str(cell)
    return False

def regex_match_with_diff(pattern: Pattern, text: str, max_diff: int) -> bool:
    """Match text to regex pattern with flexible slicing within max_diff

    Args:
        pattern: Compiled regex
        text: String to match
        max_diff: Maximum character difference allowed

    Returns:
        True if match found within max_diff
    """
    for i in range(len(text) - max_diff + 1):
        for j in range(i + 1, len(text) + 1):
            if pattern.fullmatch(text[i:j]):
                return True
    return False

def find_most_recent_date(sheet: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 150) -> Optional[datetime]:
    """Find most recent date in worksheet.

    Args:
        sheet: Excel worksheet to search
        max_rows: Maximum rows to check

    Returns:
        Most recent datetime found or None
    """
    latest_date = None
    for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for cell in row:
            if isinstance(cell, datetime): # grabs all that are dates and compares for most recent
                if not latest_date or cell > latest_date:
                    latest_date = cell
    return latest_date

def check_column_presence(workbook: openpyxl.Workbook, column_name: str, max_cols: int = 50) -> bool:
    """Check if column with certain name exists on all sheets after first.

    Args:
        workbook: Excel workbook to check
        column_name: Name of column to find
        max_cols: Maximum columns to check

    Returns:
        True if column found in any sheet
    """
    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        for column in sheet.iter_cols(min_row=1, max_row=1, max_col=max_cols, values_only=True):
            if column[0] and column_name.lower() in column[0].lower():
                return True
    return False

def check_reviewed_within_days(last_review_date: Optional[datetime],
                             days: int = 180) -> bool:
    """Check if review date is within specified number of days from today.

    Args:
        last_review_date: Date to check
        days: Number of days to check within

    Returns:
        True if date is within days
    """
    if last_review_date:
        return (datetime.now() - last_review_date).days <= days
    return False

def process_scc_file(file_path: str) -> Dict[str, Any]:
    """Process single SCC file  info.

    Args:
        file_path: Path to SCC Excel file

    Returns:
        Dict containing SCC analysis results
    """
    print(f'Performing SCC checks on {file_path}')
    workbook = read_excel(file_path)

    if workbook is None:
        print(f"Skipping file due to error: {file_path}")
        return {}

    first_sheet = workbook[workbook.sheetnames[0]]
    scm_pattern = re.compile(r'SCM\d+', re.IGNORECASE)
    guidance_pattern = re.compile(r'SCC Guidance Source', re.IGNORECASE)
    policy_procedure_pattern = re.compile(r'SCC Policy and Procedures Source', re.IGNORECASE)
    system_scope_pattern = re.compile(r'SCC System Scope', re.IGNORECASE)

    scc_name = os.path.splitext(os.path.basename(file_path))[0]
    version_match = re.search(r'_(\d{2})$', scc_name)
    version = version_match.group(1) if version_match else None
    scc_name = re.sub(r'_\d{2}$', '', scc_name)

    last_review_date = find_most_recent_date(first_sheet)
    
    scc_info = {
        'SCC': scc_name,
        'Version': version,
        'SCM Name': find_value_with_regex(first_sheet, scm_pattern, 0),
        'Last Review Date': last_review_date.isoformat() if last_review_date else None,  # Convert to ISO format string
        'SCC Guidance source presence': bool(find_value_with_regex(first_sheet, guidance_pattern, 5)),
        'SCC Policy and Procedure presence': bool(find_value_with_regex(first_sheet, policy_procedure_pattern, 5)),
        'Exception column presence': check_column_presence(workbook, 'exception'),
        'Deviation column presence': check_column_presence(workbook, 'deviation'),
        'TLA column presence': check_column_presence(workbook, 'TLA'),
        'Compliance method column presence': check_column_presence(workbook, 'method'),
        'WPS config sup doc presence': check_column_presence(workbook, 'documentation'),
        'Reviewed within 180 days': check_reviewed_within_days(last_review_date),
        'SCC System Scope Presence': bool(find_value_with_regex(first_sheet, system_scope_pattern, 3)),
        'Evidence Methods': []
    }

    return scc_info

def main() -> None:
    """Run from command line."""
    parser = argparse.ArgumentParser(description='Analyze SCC files.')
    parser.add_argument('file_path', type=str, help='Path to the Excel file')
    args = parser.parse_args()

    if not os.path.isfile(args.file_path):
        print(f"File not found: {args.file_path}")
        return

    scc_info = process_scc_file(args.file_path)
    for key, value in scc_info.items():
        print(f"{key}: {value}")

if __name__ == "__main__":
    main()
