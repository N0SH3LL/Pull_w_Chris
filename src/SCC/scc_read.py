"""Pulls information from the SCC, building out the dictionaries for progress.json. I'm sure a ton of this is deprecated now that SCC's are uniform. 
"""

import openpyxl
import re
import argparse
import os
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_excel(file_path):
    """
    Load an Excel workbook from the specified file path.
    
    Args:
    file_path (str): Path to the Excel file.
    
    Returns:
    openpyxl.Workbook or None: The loaded workbook, or None if there was an error.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook
    except PermissionError:
        logging.error(f"Permission denied: Unable to access '{file_path}'. The file may be open in another program.")
    except Exception as e:
        logging.error(f"Unexpected error while reading '{file_path}': {e}")
    return None

def find_unique_values(sheet, column_index):
    """
    Find unique values in a specific column of the sheet.
    
    Args:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to process.
    column_index (int): The index of the column to check.
    
    Returns:
    set: A set of unique values found in the column.
    """
    unique_values = set()
    for row in sheet.iter_rows(min_row=2, max_row=1000, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        if cell_value:
            unique_values.add(str(cell_value))
    return unique_values

def split_documentation_text(text):
    """
    Split a text string into separate documents based on multiple spaces or newlines.
    
    Args:
    text (str): The text to split.
    
    Returns:
    list: A list of split document names.
    """
    return re.split(r'\s{2,}|\r?\n', text)

def extract_bpers_from_cell(cell_value):
    """
    Extract BPER numbers from a cell value.
    
    Args:
    cell_value (str): The content of the cell.
    
    Returns:
    list: A list of extracted BPER numbers.
    """
    if cell_value is None:
        return []
    cell_value_str = str(cell_value)
    bper_pattern = re.compile(r'BPER\d{7}')
    return bper_pattern.findall(cell_value_str)

def process_bper_columns(sheet, row_index, exception_col, deviation_col, tla_col, scc_name):
    """
    Process BPER columns (Exception, Deviation, and TLA) for a given row.
    
    Args:
    sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet being processed.
    row_index (int): The index of the current row.
    exception_col (int): The index of the Exception column.
    deviation_col (int): The index of the Deviation column.
    tla_col (int): The index of the TLA column.
    scc_name (str): The name of the current SCC.
    
    Returns:
    list: A list of tuples containing BPER information (bper_value, is_tla, scc_name).
    """
    bper_info = []
    for col_index in [exception_col, deviation_col, tla_col]:
        if col_index is not None:
            cell_value = sheet.cell(row=row_index, column=col_index).value
            bper_values = extract_bpers_from_cell(cell_value)
            for bper_value in bper_values:
                bper_info.append((bper_value, col_index == tla_col, scc_name))
    return bper_info

def update_bper_dict(bper_dict, bper_value, scc_name, is_tla):
    """
    Update the BPER dictionary, avoiding duplicates for the same BPER and SCC combination.
    
    Args:
    bper_dict (dict): The current BPER dictionary.
    bper_value (str): The BPER number.
    scc_name (str): The name of the current SCC.
    is_tla (bool): Whether this BPER is from the TLA column.
    
    Returns:
    dict: The updated BPER dictionary.
    """
    if bper_value not in bper_dict:
        bper_dict[bper_value] = {
            'SCC': scc_name,
            'BPER name': bper_value,
            'Approval Status': '',
            'Valid to': '',
            'Gathered': False,
            'TLA': is_tla
        }
        logging.info(f"Added new BPER entry: {bper_value} for SCC {scc_name}")
    elif bper_dict[bper_value]['SCC'] == scc_name:
        if is_tla and not bper_dict[bper_value]['TLA']:
            bper_dict[bper_value]['TLA'] = True
            logging.info(f"Updated TLA status for existing BPER {bper_value} in SCC {scc_name}")
    else:
        logging.info(f"BPER {bper_value} already exists for a different SCC. Current entry not modified.")

    return bper_dict

def load_progress_data(progress_file):
    """
    Load existing progress data from progress.json.
    
    Args:
    progress_file (str): Path to the progress.json file.
    
    Returns:
    dict: The loaded progress data.
    """
    try:
        with open(progress_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        logging.warning(f"Progress file not found: {progress_file}. Starting with empty data.")
        return {'BPERs': {}, 'Documents': {}, 'Attestations': {}, 'Checks': {}}
    except json.JSONDecodeError:
        logging.error(f"Error decoding JSON from {progress_file}. Starting with empty data.")
        return {'BPERs': {}, 'Documents': {}, 'Attestations': {}, 'Checks': {}}

def process_excel_file(file_path):
    """
    Process an Excel file to extract BPERs, documents, attestations, and compliance methods.
    
    Args:
    file_path (str): Path to the Excel file to process.
    
    Returns:
    tuple: Dictionaries containing BPERs, documents, attestations, and compliance methods.
    """
    logging.info(f'Processing {file_path}')
    workbook = read_excel(file_path)

    if workbook is None:
        logging.error(f"Skipping file due to error: {file_path}")
        return {}, {}, {}, {}

    bper_dict = {}
    doc_dict = {}
    attestation_dict = {}
    method_dict = {}

    attestation_pattern = re.compile(r'(?<!\w)\d{6}(?!\w)')

    scc_name = os.path.splitext(os.path.basename(file_path))[0]
    scc_name = re.sub(r'_\d{2}$', '', scc_name).strip()

    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        exception_col = None
        deviation_col = None
        tla_col = None
        documentation_col = None
        method_col = None

        # Find relevant columns
        for col_index, column in enumerate(sheet.iter_cols(min_row=1, max_row=1, max_col=50, values_only=True), 1):
            header = str(column[0]).lower() if column[0] else ""
            if 'exception' in header:
                exception_col = col_index
            elif 'deviation' in header:
                deviation_col = col_index
            elif 'tla' in header:
                tla_col = col_index
            elif 'documentation' in header:
                documentation_col = col_index
            elif 'method' in header:
                method_col = col_index

        # Process rows
        for row_index in range(2, sheet.max_row + 1):
            # Process BPERs
            bper_info = process_bper_columns(sheet, row_index, exception_col, deviation_col, tla_col, scc_name)
            for bper_value, is_tla, bper_scc_name in bper_info:
                bper_dict = update_bper_dict(bper_dict, bper_value, bper_scc_name, is_tla)

            # Process documentation
            if documentation_col:
                doc_cell = sheet.cell(row=row_index, column=documentation_col)
                if doc_cell.value:
                    doc_names = split_documentation_text(str(doc_cell.value))
                    for doc_name in doc_names:
                        doc_name_for_comparison = doc_name.replace('\n', '').strip().upper()
                        if doc_name_for_comparison not in ['NA', 'N/A', 'NO', 'NONE']:
                            attestation_match = attestation_pattern.search(doc_name)
                            if attestation_match:
                                attestation_num = attestation_match.group()
                                if attestation_num not in attestation_dict:
                                    attestation_dict[attestation_num] = {
                                        'SCC': scc_name,
                                        'Attestation num': attestation_num,
                                        'Gathered': False,
                                        'Approval Status': '',
                                        'Valid to': ''
                                    }
                            else:
                                doc_name_final = re.sub(r'\b\d{6}\b', '', doc_name).strip()
                                if doc_name_final and doc_name_final not in doc_dict:
                                    doc_dict[doc_name_final] = {
                                        'SCC': scc_name,
                                        'Doc name': doc_name_final,
                                        'Version': '',
                                        'Last update': '',
                                        'Gathered': False
                                    }

            # Process compliance method
            if method_col:
                method_cell = sheet.cell(row=row_index, column=method_col)
                stig_id_cell = sheet.cell(row=row_index, column=1)
                if method_cell.value is not None:
                    method = str(method_cell.value)
                    stig_id = str(stig_id_cell.value)
                    method_dict[stig_id] = {
                        'SCC': scc_name,
                        'STIG ID': stig_id,
                        'Evidence Method': method,
                        'compliant': '',
                        'Gathered': False
                    }

    logging.info(f"Extracted {len(bper_dict)} BPERs, {len(doc_dict)} documents, {len(attestation_dict)} attestations, and {len(method_dict)} compliance methods from {file_path}")
    return bper_dict, doc_dict, attestation_dict, method_dict

def main():
    parser = argparse.ArgumentParser(description='Grabs BPERs and docs from an SCC.')
    parser.add_argument('file_path', type=str, help='Path to the Excel file')
    args = parser.parse_args()

    if not os.path.isfile(args.file_path):
        logging.error(f"File not found: {args.file_path}")
        return

    bper_dict, doc_dict, attestation_dict, method_dict = process_excel_file(args.file_path)

    # Print results
    print("BPER Names:")
    for bper in bper_dict:
        print(bper)
    
    print("\nDocument Names:")
    for doc in doc_dict:
        print(doc)
    
    print("\nAttestation Numbers:")
    for attestation in attestation_dict:
        print(attestation)
    
    print("\nCompliance Methods:")
    for stig_id, method_info in method_dict.items():
        print(f"STIG ID: {method_info['STIG ID']}, Method: {method_info['Evidence Method']}")

if __name__ == "__main__":
    main()
