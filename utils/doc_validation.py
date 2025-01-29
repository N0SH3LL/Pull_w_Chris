"""
DOCVALIDATION.py

This module handles updating an Excel workbook that tracks and validates various document statuses 
across different tabs for Security Control Checklists (SCCs). It maintains information about SCCs, 
SCMs, documents, BPERs, and attestations.

The module uses openpyxl to manipulate Excel files and handles data from a progress.json file
that contains the current state of all tracked items.
"""

import openpyxl
import json
from datetime import datetime

def update_document_validation(progress_file, template_path):
    """
    Main function to update all tabs in the document validation workbook.
    
    Args:
        progress_file (str): Path to the progress.json file containing current state
        template_path (str): Path to the Excel template to be updated
    """
    # Load the progress data from JSON
    with open(progress_file, 'r') as file:
        progress_data = json.load(file)
    
    # Load the Excel workbook for updating
    workbook = openpyxl.load_workbook(template_path)
    
    # Update each tab in the workbook
    update_sccs_tab(workbook, progress_data)
    update_scc_scm_tab(workbook, progress_data)
    update_scc_documents_tab(workbook, progress_data)
    update_scc_bper_tab(workbook, progress_data)
    update_scc_attestation_tab(workbook, progress_data)
    
    # Save changes back to the workbook
    workbook.save(template_path)

def update_sccs_tab(workbook, progress_data):
    """
    Update the SCCs tab with general SCC information and status indicators.
    
    Args:
        workbook: Open workbook object
        progress_data (dict): Data from progress.json containing SCC information
    """
    sheet = workbook['SCC\'s']
    row = 2  # Start from row 2 (after headers)
    
    for scc_path, scc_data in progress_data['SCC'].items():
        # Update basic SCC information
        sheet.cell(row=row, column=2, value=scc_data.get('SCC', ''))
        sheet.cell(row=row, column=3, value=scc_data.get('Version', ''))
        
        # Handle last review date formatting
        last_review_date = scc_data.get('Last Review Date', '')
        if last_review_date:
            if isinstance(last_review_date, str):
                last_review_date = last_review_date.replace('T00:00:00', '')
            elif isinstance(last_review_date, datetime):
                last_review_date = last_review_date.strftime('%Y-%m-%d')
            sheet.cell(row=row, column=4, value=last_review_date)
        
        # Update SCM indicator
        sheet.cell(row=row, column=6, value='X' if scc_data.get('SCM Name') else '')
        
        # Map of column numbers to boolean fields
        boolean_columns = {
            7: 'SCC System Scope Presence',
            8: 'SCC Policy and Procedure presence',
            9: 'Compliance method column presence',
            10: 'Exception column presence',
            11: 'Deviation column presence',
            12: 'TLA column presence',
            13: 'Compliance method column presence',
            14: 'WPS config sup doc presence'
        }
        
        # Update boolean indicators
        for col, key in boolean_columns.items():
            value = scc_data.get(key, False)
            sheet.cell(row=row, column=col, value='X' if value else '')
        
        row += 1

def update_scc_scm_tab(workbook, progress_data):
    """
    Update the SCC-SCM tab with mappings between SCCs and their associated SCMs.
    
    Args:
        workbook: Open workbook object
        progress_data (dict): Data from progress.json containing SCC and SCM relationships
        
    This function populates a two-column mapping showing which SCMs are associated
    with each SCC in the system.
    """
    sheet = workbook['SCC-SCM']
    row = 2  # Start from row 2 (after headers)
    
    for scc_path, scc_data in progress_data['SCC'].items():
        # Write SCC name
        sheet.cell(row=row, column=1, value=scc_data.get('SCC', ''))
        
        # Write associated SCM name if it exists
        scm_name = scc_data.get('SCM Name', '')
        if scm_name:  # Only write SCM Name if it's not false or empty
            sheet.cell(row=row, column=2, value=scm_name)
        row += 1  

def update_scc_documents_tab(workbook, progress_data):
    """
    Update the SCC-Documents tab with information about supporting documentation.
    
    Args:
        workbook: Open workbook object
        progress_data (dict): Data from progress.json containing document information
        
    Tracks document details including:
        - Associated SCC
        - Document name
        - Version information
        - Gathering status
        - Timestamps for gathering and updates
    """
    sheet = workbook['SCC-Documents']
    row = 2  # Start from row 2 (after headers)
    
    for doc_name, doc_data_list in progress_data['Documents'].items():
        for doc_data in doc_data_list:
            # Write document information
            sheet.cell(row=row, column=1, value=doc_data.get('SCC', ''))
            sheet.cell(row=row, column=2, value=doc_data.get('Doc name', ''))
            sheet.cell(row=row, column=3, value=doc_data.get('Version', ''))
            
            # Update gathering status and timestamps
            sheet.cell(row=row, column=4, value='X' if doc_data.get('Gathered') else '')
            sheet.cell(row=row, column=5, value=doc_data.get('Gathered timestamp', ''))
            sheet.cell(row=row, column=6, value=doc_data.get('Last update', ''))
            row += 1

def update_scc_bper_tab(workbook, progress_data):
    """
    Update the SCC-BPER tab with Business Process Exception Request information.
    
    Args:
        workbook: Open workbook object
        progress_data (dict): Data from progress.json containing BPER information
        
    Tracks BPER details including:
        - Associated SCC
        - BPER identifier
        - Gathering status
        - Approval status
        - Validity dates
    """
    sheet = workbook['SCC-BPER']
    row = 2  # Start from row 2 (after headers)

    for bper_name, bper_data_list in progress_data['BPERs'].items():
        for bper_data in bper_data_list:
            # Write BPER information
            sheet.cell(row=row, column=1, value=bper_data.get('SCC', ''))
            sheet.cell(row=row, column=2, value=bper_name)
            
            # Update gathering and approval statuses
            sheet.cell(row=row, column=3, value='X' if bper_data.get('Gathered') else '')
            gathered_timestamp = bper_data.get('Gathered timestamp', '')
            sheet.cell(row=row, column=4, value=gathered_timestamp.split()[0] if gathered_timestamp else '')
            sheet.cell(row=row, column=5, value='X' if bper_data.get('Approval Status') == 'Approved' else '')
            sheet.cell(row=row, column=6, value=bper_data.get('Valid to', ''))
            row += 1

def update_scc_attestation_tab(workbook, progress_data):
    """
    Update the SCC-Attestation tab with attestation tracking information.
    
    Args:
        workbook: Open workbook object
        progress_data (dict): Data from progress.json containing attestation information
        
    Tracks attestation details including:
        - Associated SCC
        - Attestation number
        - Gathering status and timestamps
        - Approval status
        - Validity dates
        - Review dates
        - Overall status
        
    Note: This function expects attestation data to follow a specific format where
    approval status of 'approve open' is treated as a special case for status indicators.
    """
    sheet = workbook['SCC-Attestation']
    row = 2  # Start from row 2 (after headers)

    for attestation_num, attestation_data_list in progress_data['Attestations'].items():
        for attestation_data in attestation_data_list:
            # Basic attestation information
            sheet.cell(row=row, column=1, value=attestation_data.get('SCC', ''))
            sheet.cell(row=row, column=2, value=attestation_num)
            
            # Gathering status and timestamps
            sheet.cell(row=row, column=3, value='X' if attestation_data.get('Gathered') else '')
            sheet.cell(row=row, column=4, value=attestation_data.get('Gathered timestamp', ''))
            
            # Approval and validity information
            sheet.cell(row=row, column=5, value='X' if attestation_data.get('Approval Status', '') == 'approve open' else '')
            sheet.cell(row=row, column=6, value=attestation_data.get('Valid to', ''))
            
            # Review and overall status
            sheet.cell(row=row, column=7, value=attestation_data.get('Review Date'))
            sheet.cell(row=row, column=8, value=attestation_data.get('Overall Status', ''))
            row += 1
